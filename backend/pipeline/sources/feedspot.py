"""
feedspot.py — Normalizer for Feedspot XLSX/CSV files.

Handles the BloggerOutreach FullDB export format:
  - Multiple XLSX files, each with one or more named sheets
  - Two schema variants detected per-sheet from header col0:
      V1 (col0='Id'):    18 cols, Description at col17, no Domain Authority
      V2 (col0='Sr.no'): 19 cols, Domain Authority at col17, Description at col18

Actual files confirmed in ./feedspot/:
  _BloggerOutreach_FullDB_Blogs(2_2)_...xlsx        → 'Blog Media List (22)'    V1  ~83k rows
  _BloggerOutreach_FullDB_Magazines_...xlsx          → 'Magazine Media List'     V1  ~14k rows
  _BloggerOutreach_FullDB_Podcasts_...xlsx           → 'Podcast Media List'      V2  ~85k rows
  _BloggerOutreach_Youtube_FullDB_70K_...xlsx        → 'YouTube Media List'      V1  ~76k rows
  _BloggerOutreach_Youtube_FullDB_70K_...copy.xlsx   → duplicate of above (dedup handles it)

Sub-author rows: where Site URL is empty but Author Name is present, we attach
the person to the last seen publisher via a HAS_AUTHOR edge.
"""
from __future__ import annotations

import csv
import logging
import os
import sys
from pathlib import Path
from typing import Iterator, List, Optional, Tuple

import openpyxl

from backend.config import FEEDSPOT_DIR, FEEDSPOT_EXPECTED_SHEETS
from backend.models import PersonNode, PublisherNode, EdgeRecord
from backend.utils import (
    clean_text, clean_url, clean_handle, person_id, publisher_id,
    parse_follower_count, parse_domain_authority, compute_reach_score,
    infer_topic_cluster, infer_category_type, compute_confidence,
    compute_warmth, utcnow_str, safe_get,
)

csv.field_size_limit(10_000_000)

logger = logging.getLogger(__name__)

# ── Schema detection ──────────────────────────────────────────────────────────

def _detect_schema(header_row: tuple) -> int:
    """
    Returns 1 (V1) or 2 (V2) based on col0 of the header row.
    V1: col0 = 'Id'
    V2: col0 = 'Sr.no'
    Falls back to 1 if unrecognised.
    """
    col0 = str(header_row[0]).strip().lower() if header_row else ""
    if col0 == "sr.no":
        return 2
    return 1


# ── Row parsers ───────────────────────────────────────────────────────────────

def _parse_row_v1(
    row: tuple,
    category: str,
    now: str,
    last_pub_id: Optional[str],
) -> Tuple[Optional[PublisherNode], Optional[PersonNode], Optional[EdgeRecord]]:
    """
    V1 columns:
      0=Id  1=Category  2=Site URL  3=Site Name  4=Author Name
      5=Primary Email  6=Author Designation  7=Author Twitter  8=Author Linkedin
      9=Notes  10=Location  11=FB Followers  12=FB Url  13=TW Followers
      14=TW Url  15=IG Followers  16=IG URL  17=Description
    """
    site_url    = clean_url(safe_get(row, 2))
    site_name   = clean_text(safe_get(row, 3))
    author_name = clean_text(safe_get(row, 4))
    email       = clean_text(safe_get(row, 5)).lower() if safe_get(row, 5) else None
    designation = clean_text(safe_get(row, 6))
    tw_handle   = clean_handle(safe_get(row, 7))
    linkedin_url = clean_url(safe_get(row, 8))
    location    = clean_text(safe_get(row, 10))
    fb_followers = parse_follower_count(safe_get(row, 11))
    fb_url      = clean_url(safe_get(row, 12))
    tw_followers = parse_follower_count(safe_get(row, 13))
    tw_url      = clean_url(safe_get(row, 14))
    ig_followers = parse_follower_count(safe_get(row, 15))
    ig_url      = clean_url(safe_get(row, 16))
    description = clean_text(safe_get(row, 17))

    # Sub-author row: no Site URL/Name but has Author Name
    if not site_url and not site_name:
        if author_name and last_pub_id:
            person, edge = _build_person(
                last_pub_id, author_name, email, tw_handle, linkedin_url,
                location, designation, category, now
            )
            return None, person, edge
        return None, None, None

    pub_id = publisher_id(site_url or None, site_name or None)
    reach = compute_reach_score(fb_followers, tw_followers, ig_followers)
    topic = infer_topic_cluster(f"{site_name} {description} {category}")
    cat_type = infer_category_type(category, site_name, description)

    pub_fields = [site_url, site_name, description, location,
                  fb_followers, tw_followers, ig_followers, email]
    pub_conf = compute_confidence(1, pub_fields)

    publisher = PublisherNode(
        id=pub_id,
        name=site_name or site_url,
        source="feedspot",
        site_url=site_url or None,
        category=category,
        category_type=cat_type,
        topic_cluster=topic,
        description=description or None,
        location=location or None,
        domain_authority=0,
        fb_followers=fb_followers,
        tw_followers=tw_followers,
        ig_followers=ig_followers,
        reach_score=reach,
        email=email or None,
        fb_url=fb_url or None,
        tw_url=tw_url or None,
        ig_url=ig_url or None,
        confidence_score=pub_conf,
        ingested_at=now,
    )

    person: Optional[PersonNode] = None
    edge: Optional[EdgeRecord] = None
    if author_name:
        person, edge = _build_person(
            pub_id, author_name, email, tw_handle, linkedin_url,
            location, designation, category, now
        )

    return publisher, person, edge


def _parse_row_v2(
    row: tuple,
    category: str,
    now: str,
    last_pub_id: Optional[str],
) -> Tuple[Optional[PublisherNode], Optional[PersonNode], Optional[EdgeRecord]]:
    """
    V2 columns:
      0=Sr.no  1=Category  2=Site URL  3=Site Name  4=Author Name
      5=Primary Email  6=Designation  7=Twitter  8=Linkedin
      9=Notes  10=Location  11=FB Followers  12=FB Url  13=TW Followers
      14=TW Url  15=IG Followers  16=IG URL  17=Domain Authority  18=Description
    """
    site_url    = clean_url(safe_get(row, 2))
    site_name   = clean_text(safe_get(row, 3))
    author_name = clean_text(safe_get(row, 4))
    email       = clean_text(safe_get(row, 5)).lower() if safe_get(row, 5) else None
    designation = clean_text(safe_get(row, 6))
    tw_handle   = clean_handle(safe_get(row, 7))
    linkedin_url = clean_url(safe_get(row, 8))
    location    = clean_text(safe_get(row, 10))
    fb_followers = parse_follower_count(safe_get(row, 11))
    fb_url      = clean_url(safe_get(row, 12))
    tw_followers = parse_follower_count(safe_get(row, 13))
    tw_url      = clean_url(safe_get(row, 14))
    ig_followers = parse_follower_count(safe_get(row, 15))
    ig_url      = clean_url(safe_get(row, 16))
    domain_auth = parse_domain_authority(safe_get(row, 17))
    description = clean_text(safe_get(row, 18))

    # Sub-author row
    if not site_url and not site_name:
        if author_name and last_pub_id:
            person, edge = _build_person(
                last_pub_id, author_name, email, tw_handle, linkedin_url,
                location, designation, category, now
            )
            return None, person, edge
        return None, None, None

    pub_id = publisher_id(site_url or None, site_name or None)
    reach = compute_reach_score(fb_followers, tw_followers, ig_followers)
    topic = infer_topic_cluster(f"{site_name} {description} {category}")
    cat_type = infer_category_type(category, site_name, description)

    pub_fields = [site_url, site_name, description, location,
                  fb_followers, tw_followers, ig_followers, email, domain_auth]
    pub_conf = compute_confidence(1, pub_fields)

    publisher = PublisherNode(
        id=pub_id,
        name=site_name or site_url,
        source="feedspot",
        site_url=site_url or None,
        category=category,
        category_type=cat_type,
        topic_cluster=topic,
        description=description or None,
        location=location or None,
        domain_authority=domain_auth,
        fb_followers=fb_followers,
        tw_followers=tw_followers,
        ig_followers=ig_followers,
        reach_score=reach,
        email=email or None,
        fb_url=fb_url or None,
        tw_url=tw_url or None,
        ig_url=ig_url or None,
        confidence_score=pub_conf,
        ingested_at=now,
    )

    person: Optional[PersonNode] = None
    edge: Optional[EdgeRecord] = None
    if author_name:
        person, edge = _build_person(
            pub_id, author_name, email, tw_handle, linkedin_url,
            location, designation, category, now
        )

    return publisher, person, edge


def _build_person(
    pub_id: str,
    author_name: str,
    email: Optional[str],
    tw_handle: str,
    linkedin_url: str,
    location: str,
    designation: str,
    category: str,
    now: str,
) -> Tuple[PersonNode, EdgeRecord]:
    p_id = person_id(
        x_handle=tw_handle or None,
        email=email or None,
        name=author_name,
        src="feedspot",
    )
    per_fields = [author_name, email, tw_handle, linkedin_url, location, designation]
    per_conf = compute_confidence(1, per_fields)
    topic = infer_topic_cluster(f"{author_name} {designation} {category}")

    person = PersonNode(
        id=p_id,
        name=author_name,
        source=["feedspot"],
        x_handle=tw_handle or None,
        x_url=f"https://x.com/{tw_handle}" if tw_handle else None,
        email=email or None,
        linkedin_url=linkedin_url or None,
        location=location or None,
        warmth_score=compute_warmth(["feedspot"]),
        confidence_score=per_conf,
        ingested_at=now,
        topic_cluster=topic,
    )
    edge = EdgeRecord(
        source_id=pub_id,
        target_id=p_id,
        rel_type="HAS_AUTHOR",
        weight=1.0,
    )
    return person, edge


# ── XLSX sheet parser ─────────────────────────────────────────────────────────

def _parse_worksheet(
    ws,
    sheet_name: str,
    filepath: str,
) -> Iterator[Tuple[Optional[PublisherNode], Optional[PersonNode], Optional[EdgeRecord]]]:
    """Parse a single openpyxl worksheet, yielding (pub, person, edge) tuples."""
    now = utcnow_str()
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        return

    schema = _detect_schema(header)
    last_pub_id: Optional[str] = None
    row_count = 0

    for row in rows_iter:
        # Skip completely empty rows
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        row_count += 1
        category = clean_text(safe_get(row, 1)) or "General"

        try:
            if schema == 2:
                pub, person, edge = _parse_row_v2(row, category, now, last_pub_id)
            else:
                pub, person, edge = _parse_row_v1(row, category, now, last_pub_id)
        except Exception as exc:
            logger.debug("Skipping row in %s/%s: %s", filepath, sheet_name, exc)
            continue

        if pub is not None:
            last_pub_id = pub.id

        yield pub, person, edge


# ── CSV sheet parser (legacy fallback) ───────────────────────────────────────

def _parse_csv_file(
    filepath: str,
) -> Iterator[Tuple[Optional[PublisherNode], Optional[PersonNode], Optional[EdgeRecord]]]:
    """Parse a legacy Feedspot CSV file."""
    now = utcnow_str()
    last_pub_id: Optional[str] = None

    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            return

        schema = _detect_schema(tuple(header))

        for row in reader:
            if not row or all(c.strip() == "" for c in row):
                continue
            category = clean_text(safe_get(row, 1)) or "General"
            try:
                if schema == 2:
                    pub, person, edge = _parse_row_v2(tuple(row), category, now, last_pub_id)
                else:
                    pub, person, edge = _parse_row_v1(tuple(row), category, now, last_pub_id)
            except Exception as exc:
                logger.debug("Skipping CSV row in %s: %s", filepath, exc)
                continue

            if pub is not None:
                last_pub_id = pub.id

            yield pub, person, edge


# ── Public entry point ────────────────────────────────────────────────────────

def parse_feedspot_file(
    filepath: str,
) -> Iterator[Tuple[Optional[PublisherNode], Optional[PersonNode], Optional[EdgeRecord]]]:
    """
    Parse a single Feedspot file (XLSX or CSV).
    For XLSX: iterates all sheets.
    For CSV: reads as before.
    Yields (publisher_or_None, person_or_None, edge_or_None).
    """
    ext = Path(filepath).suffix.lower()

    if ext == ".xlsx":
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as exc:
            logger.warning("Cannot open %s: %s", filepath, exc)
            return

        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                yield from _parse_worksheet(ws, sheet_name, filepath)
            except Exception as exc:
                logger.warning("Skipping sheet %s in %s: %s", sheet_name, filepath, exc)
                continue

        wb.close()

    else:
        yield from _parse_csv_file(filepath)


# ── CLI test harness ──────────────────────────────────────────────────────────

def _run_test() -> None:
    """
    python -m backend.pipeline.sources.feedspot --test

    Prints: file count, sheet count, total rows found, schema variant
    distribution, and 3 sample PublisherNode records per media type.
    """
    import json
    from collections import defaultdict

    feedspot_dir = FEEDSPOT_DIR
    files = sorted(
        p for p in Path(feedspot_dir).iterdir()
        if p.suffix.lower() in (".xlsx", ".csv") and not p.name.startswith(".")
    )

    print(f"\n{'='*60}")
    print(f"FEEDSPOT TEST — {feedspot_dir}")
    print(f"Files found: {len(files)}")
    for f in files:
        print(f"  {f.name}")

    # Count sheets
    total_sheets = 0
    for f in files:
        if f.suffix.lower() == ".xlsx":
            try:
                wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
                total_sheets += len(wb.sheetnames)
                wb.close()
            except Exception:
                pass
        else:
            total_sheets += 1

    print(f"Total sheets: {total_sheets}")
    print(f"{'='*60}\n")

    counts: dict = defaultdict(int)
    schema_dist: dict = defaultdict(int)
    samples: dict = defaultdict(list)
    type_map = {
        "Blog": "feedspot_blogs",
        "Podcast": "feedspot_podcasts",
        "YouTube": "feedspot_youtube",
        "Magazine": "feedspot_magazines",
        "Newsletter": "feedspot_newsletters",
    }

    for filepath in files:
        ext = Path(filepath).suffix.lower()
        if ext == ".xlsx":
            try:
                wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
            except Exception as e:
                print(f"SKIP {filepath.name}: {e}")
                continue

            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    rows_iter = ws.iter_rows(values_only=True)
                    header = next(rows_iter, None)
                    if header is None:
                        continue
                    schema = _detect_schema(header)
                    schema_dist[f"V{schema}"] += 1

                    for row in rows_iter:
                        if not row or all(v is None or str(v).strip() == "" for v in row):
                            continue
                        site_url = clean_url(safe_get(row, 2))
                        site_name = clean_text(safe_get(row, 3))
                        category = clean_text(safe_get(row, 1)) or "General"
                        if not site_url and not site_name:
                            continue
                        cat_type = infer_category_type(category, site_name, "")
                        key = type_map.get(cat_type, "feedspot_other")
                        counts[key] += 1
                        if len(samples[cat_type]) < 3:
                            samples[cat_type].append({
                                "name": site_name,
                                "url": site_url,
                                "category": category,
                                "category_type": cat_type,
                                "sheet": sheet_name,
                                "schema": f"V{schema}",
                            })
                except Exception as e:
                    print(f"  SKIP sheet {sheet_name}: {e}")
                    continue

            wb.close()

    print("Schema distribution:")
    for k, v in sorted(schema_dist.items()):
        print(f"  {k}: {v} sheet(s)")

    print("\nRow counts by media type:")
    total = 0
    for k, v in sorted(counts.items()):
        print(f"  {k:<30s}: {v:>8,}")
        total += v
    print(f"  {'TOTAL':<30s}: {total:>8,}")

    print("\nSample records:")
    for cat_type, recs in sorted(samples.items()):
        print(f"\n  [{cat_type}]")
        for r in recs:
            print(f"    {r['name']!r:<45s} {r['url'][:50]}")
            print(f"      category={r['category']!r}  schema={r['schema']}  sheet={r['sheet']!r}")

    print(f"\n{'='*60}\n")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Feedspot normalizer test")
    parser.add_argument("--test", action="store_true", help="Run test/stats mode")
    args = parser.parse_args()
    if args.test:
        _run_test()
    else:
        parser.print_help()
